"""
Excel Logging System for Trading Data
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import logging
from typing import List, Dict, Optional
from config import *
from portfolio_manager import Trade

logger = logging.getLogger(__name__)

class ExcelLogger:
    def __init__(self, excel_file: str = EXCEL_FILE):
        self.excel_file = excel_file
        self.columns = [
            "Date", "Ticker", "Market Direction", "Strategy", "Strategy Version",
            "Entry Time", "Exit Time", "Entry Price (intended)", "Entry Fill Price",
            "Exit Price (intended)", "Exit Fill Price", "Stop Loss Price", "Shares",
            "Total Price", "Account Size ($)", "$ Risked (per trade)", "% of Portfolio Risked",
            "Gross P/L ($)", "% Return on Trade", "R Multiple", "Trade Duration (min)",
            "Win/Loss", "Setup Quality (1–5)", "Order ID", "Commission", "Slippage",
            "ATR at Entry", "Entry Signal", "Exit Signal", "Notes", "Closing Notes",
            "Cumulative P/L"
        ]
        
        self._initialize_excel_file()
    
    def _initialize_excel_file(self):
        """Initialize Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.excel_file):
            try:
                # Create new workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Trading Log"
                
                # Add headers
                for col, header in enumerate(self.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Set column widths
                column_widths = {
                    'A': 12, 'B': 10, 'C': 15, 'D': 12, 'E': 15,
                    'F': 20, 'G': 20, 'H': 18, 'I': 15, 'J': 18,
                    'K': 15, 'L': 15, 'M': 8, 'N': 12, 'O': 15,
                    'P': 18, 'Q': 20, 'R': 12, 'S': 18, 'T': 12,
                    'U': 18, 'V': 10, 'W': 18, 'X': 12, 'Y': 12,
                    'Z': 10, 'AA': 12, 'AB': 15, 'AC': 15, 'AD': 30,
                    'AE': 30, 'AF': 15
                }
                
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # Create summary sheet
                summary_ws = wb.create_sheet("Summary")
                self._create_summary_sheet(summary_ws)
                
                # Create performance sheet
                performance_ws = wb.create_sheet("Performance")
                self._create_performance_sheet(performance_ws)
                
                wb.save(self.excel_file)
                logger.info(f"Created new Excel file: {self.excel_file}")
                
            except Exception as e:
                logger.error(f"Error creating Excel file: {e}")
    
    def _create_summary_sheet(self, ws):
        """Create summary sheet with key metrics"""
        headers = [
            "Metric", "Value", "Description"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Add summary data placeholders
        summary_data = [
            ["Total Trades", "", "Total number of completed trades"],
            ["Winning Trades", "", "Number of profitable trades"],
            ["Losing Trades", "", "Number of losing trades"],
            ["Win Rate (%)", "", "Percentage of winning trades"],
            ["Total Profit ($)", "", "Total profit from winning trades"],
            ["Total Loss ($)", "", "Total loss from losing trades"],
            ["Net Profit ($)", "", "Net profit/loss"],
            ["Average Win ($)", "", "Average profit per winning trade"],
            ["Average Loss ($)", "", "Average loss per losing trade"],
            ["Profit Factor", "", "Total profit / Total loss"],
            ["Max Drawdown (%)", "", "Maximum peak-to-trough decline"],
            ["Sharpe Ratio", "", "Risk-adjusted return measure"],
            ["Last Updated", "", "Last time data was updated"]
        ]
        
        for row, data in enumerate(summary_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
    
    def _create_performance_sheet(self, ws):
        """Create performance tracking sheet"""
        headers = [
            "Date", "Total Trades", "Winning Trades", "Losing Trades", 
            "Win Rate (%)", "Daily P/L ($)", "Cumulative P/L ($)", 
            "Drawdown (%)", "Trades by Strategy"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Set column widths
        column_widths = {'A': 12, 'B': 12, 'C': 15, 'D': 15, 'E': 12, 
                        'F': 15, 'G': 15, 'H': 12, 'I': 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def log_trade(self, trade: Trade, market_direction: str = "", 
                  atr_at_entry: float = 0.0, entry_signal: str = ""):
        """Log a completed trade to Excel"""
        try:
            # Read existing data
            df = pd.read_excel(self.excel_file, sheet_name="Trading Log")
            
            # Create new row
            new_row = {
                "Date": trade.date,
                "Ticker": trade.ticker,
                "Market Direction": market_direction,
                "Strategy": trade.strategy,
                "Strategy Version": trade.strategy_version,
                "Entry Time": trade.entry_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Exit Time": trade.exit_time.strftime("%Y-%m-%d %H:%M:%S") if trade.exit_time else "",
                "Entry Price (intended)": trade.entry_price_intended,
                "Entry Fill Price": trade.entry_fill_price,
                "Exit Price (intended)": trade.exit_price_intended or "",
                "Exit Fill Price": trade.exit_fill_price or "",
                "Stop Loss Price": trade.stop_loss_price,
                "Shares": trade.shares,
                "Total Price": trade.total_price,
                "Account Size ($)": trade.account_size,
                "$ Risked (per trade)": trade.risk_amount,
                "% of Portfolio Risked": trade.portfolio_risk_pct,
                "Gross P/L ($)": trade.gross_pl or 0,
                "% Return on Trade": trade.return_pct or 0,
                "R Multiple": trade.r_multiple or 0,
                "Trade Duration (min)": trade.trade_duration_min or 0,
                "Win/Loss": trade.win_loss or "",
                "Setup Quality (1–5)": trade.setup_quality,
                "Order ID": trade.order_id,
                "Commission": trade.commission,
                "Slippage": trade.slippage,
                "ATR at Entry": atr_at_entry,
                "Entry Signal": entry_signal,
                "Exit Signal": trade.exit_signal,
                "Notes": trade.notes,
                "Closing Notes": trade.closing_notes,
                "Cumulative P/L": trade.cumulative_pl
            }
            
            # Append new row
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            
            # Write back to Excel
            self._write_to_excel(df)
            
            logger.info(f"Logged trade {trade.order_id} to Excel")
            
        except Exception as e:
            logger.error(f"Error logging trade to Excel: {e}")
    
    def log_multiple_trades(self, trades: List[Trade], market_directions: Dict[str, str] = None,
                           atr_data: Dict[str, float] = None, entry_signals: Dict[str, str] = None):
        """Log multiple trades at once"""
        try:
            # Read existing data
            df = pd.read_excel(self.excel_file, sheet_name="Trading Log")
            
            new_rows = []
            for trade in trades:
                market_direction = market_directions.get(trade.ticker, "") if market_directions else ""
                atr_at_entry = atr_data.get(trade.ticker, 0.0) if atr_data else 0.0
                entry_signal = entry_signals.get(trade.ticker, "") if entry_signals else ""
                
                new_row = {
                    "Date": trade.date,
                    "Ticker": trade.ticker,
                    "Market Direction": market_direction,
                    "Strategy": trade.strategy,
                    "Strategy Version": trade.strategy_version,
                    "Entry Time": trade.entry_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "Exit Time": trade.exit_time.strftime("%Y-%m-%d %H:%M:%S") if trade.exit_time else "",
                    "Entry Price (intended)": trade.entry_price_intended,
                    "Entry Fill Price": trade.entry_fill_price,
                    "Exit Price (intended)": trade.exit_price_intended or "",
                    "Exit Fill Price": trade.exit_fill_price or "",
                    "Stop Loss Price": trade.stop_loss_price,
                    "Shares": trade.shares,
                    "Total Price": trade.total_price,
                    "Account Size ($)": trade.account_size,
                    "$ Risked (per trade)": trade.risk_amount,
                    "% of Portfolio Risked": trade.portfolio_risk_pct,
                    "Gross P/L ($)": trade.gross_pl or 0,
                    "% Return on Trade": trade.return_pct or 0,
                    "R Multiple": trade.r_multiple or 0,
                    "Trade Duration (min)": trade.trade_duration_min or 0,
                    "Win/Loss": trade.win_loss or "",
                    "Setup Quality (1–5)": trade.setup_quality,
                    "Order ID": trade.order_id,
                    "Commission": trade.commission,
                    "Slippage": trade.slippage,
                    "ATR at Entry": atr_at_entry,
                    "Entry Signal": entry_signal,
                    "Exit Signal": trade.exit_signal,
                    "Notes": trade.notes,
                    "Closing Notes": trade.closing_notes,
                    "Cumulative P/L": trade.cumulative_pl
                }
                new_rows.append(new_row)
            
            # Append new rows
            df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
            
            # Write back to Excel
            self._write_to_excel(df)
            
            logger.info(f"Logged {len(trades)} trades to Excel")
            
        except Exception as e:
            logger.error(f"Error logging multiple trades to Excel: {e}")
    
    def _write_to_excel(self, df: pd.DataFrame):
        """Write DataFrame to Excel with formatting"""
        try:
            with pd.ExcelWriter(self.excel_file, engine='openpyxl', mode='w') as writer:
                # Write main trading log
                df.to_excel(writer, sheet_name='Trading Log', index=False)
                
                # Get workbook and worksheet for formatting
                wb = writer.book
                ws = wb['Trading Log']
                
                # Format headers
                for col in range(1, len(self.columns) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Format data rows
                for row in range(2, len(df) + 2):
                    for col in range(1, len(self.columns) + 1):
                        cell = ws.cell(row=row, column=col)
                        
                        # Color code win/loss
                        if col == self.columns.index("Win/Loss") + 1:  # Win/Loss column
                            if cell.value == "Win":
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif cell.value == "Loss":
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        
                        # Color code P/L
                        elif col == self.columns.index("Gross P/L ($)") + 1:  # P/L column
                            if isinstance(cell.value, (int, float)):
                                if cell.value > 0:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                elif cell.value < 0:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Recreate summary and performance sheets
                summary_ws = wb.create_sheet("Summary")
                performance_ws = wb.create_sheet("Performance")
                self._create_summary_sheet(summary_ws)
                self._create_performance_sheet(performance_ws)
                
                # Update summary with current data
                self._update_summary_sheet(summary_ws, df)
                
        except Exception as e:
            logger.error(f"Error writing to Excel: {e}")
    
    def _update_summary_sheet(self, ws, df: pd.DataFrame):
        """Update summary sheet with current statistics"""
        try:
            if df.empty:
                return
            
            # Calculate metrics
            total_trades = len(df)
            winning_trades = len(df[df['Win/Loss'] == 'Win'])
            losing_trades = len(df[df['Win/Loss'] == 'Loss'])
            win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
            
            total_profit = df[df['Gross P/L ($)'] > 0]['Gross P/L ($)'].sum() if winning_trades > 0 else 0
            total_loss = abs(df[df['Gross P/L ($)'] < 0]['Gross P/L ($)'].sum()) if losing_trades > 0 else 0
            net_profit = total_profit - total_loss
            
            avg_win = total_profit / winning_trades if winning_trades > 0 else 0
            avg_loss = total_loss / losing_trades if losing_trades > 0 else 0
            profit_factor = total_profit / total_loss if total_loss > 0 else float('inf')
            
            # Calculate drawdown
            df['Cumulative P/L'] = df['Cumulative P/L'].fillna(0)
            peak = df['Cumulative P/L'].expanding().max()
            drawdown = (df['Cumulative P/L'] - peak) / (peak + INITIAL_PORTFOLIO_VALUE) * 100
            max_drawdown = drawdown.min()
            
            # Update summary data
            summary_data = [
                ["Total Trades", total_trades, "Total number of completed trades"],
                ["Winning Trades", winning_trades, "Number of profitable trades"],
                ["Losing Trades", losing_trades, "Number of losing trades"],
                ["Win Rate (%)", f"{win_rate:.2f}", "Percentage of winning trades"],
                ["Total Profit ($)", f"{total_profit:.2f}", "Total profit from winning trades"],
                ["Total Loss ($)", f"{total_loss:.2f}", "Total loss from losing trades"],
                ["Net Profit ($)", f"{net_profit:.2f}", "Net profit/loss"],
                ["Average Win ($)", f"{avg_win:.2f}", "Average profit per winning trade"],
                ["Average Loss ($)", f"{avg_loss:.2f}", "Average loss per losing trade"],
                ["Profit Factor", f"{profit_factor:.2f}", "Total profit / Total loss"],
                ["Max Drawdown (%)", f"{max_drawdown:.2f}", "Maximum peak-to-trough decline"],
                ["Sharpe Ratio", "TBD", "Risk-adjusted return measure"],
                ["Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Last time data was updated"]
            ]
            
            for row, data in enumerate(summary_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                    
        except Exception as e:
            logger.error(f"Error updating summary sheet: {e}")
    
    def get_trading_history(self) -> pd.DataFrame:
        """Get all trading history from Excel"""
        try:
            return pd.read_excel(self.excel_file, sheet_name="Trading Log")
        except Exception as e:
            logger.error(f"Error reading trading history: {e}")
            return pd.DataFrame()
    
    def export_to_csv(self, filename: str = None):
        """Export trading log to CSV"""
        try:
            if not filename:
                filename = f"trading_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            df = self.get_trading_history()
            df.to_csv(filename, index=False)
            logger.info(f"Exported trading log to {filename}")
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
